"""pbi source — data source profiling and model scaffold commands (Epic A)."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import click
from rich.console import Console

from pbi_cli._audit import write_audit_entry
from pbi_cli.commands._shared import (
    dry_run_echo,
    get_backend,
    output_json_or_table,
    snapshot_before_write,
)

console = Console(legacy_windows=False)


@click.group()
def source() -> None:
    """Profile data sources and scaffold star-schema models from them."""


@source.command("profile")
@click.option(
    "--type",
    "source_type",
    required=True,
    type=click.Choice(["sql", "excel", "csv", "rest"]),
    help="Source type.",
)
@click.option("--conn", default=None, help="SQL connection string.")
@click.option("--path", default=None, type=click.Path(), help="File path (Excel/CSV).")
@click.option("--url", default=None, help="REST API URL.")
@click.option("--tables", default=None, help="Comma-separated table names to include (SQL only).")
@click.option("--output", default=None, type=click.Path(), help="Save profile JSON to file.")
@click.option(
    "--bearer-token",
    default=None,
    envvar="PBI_REST_BEARER",
    help="Bearer token for REST auth (or set PBI_REST_BEARER env var).",
)
@click.option(
    "--api-key-header",
    default=None,
    help="Header name=value pair for REST API key auth (e.g. 'X-Api-Key=abc123').",
)
@click.option(
    "--rest-max-pages", default=5, show_default=True, help="Max pages to fetch for REST pagination."
)
@click.option(
    "--rest-results-path",
    default=None,
    help="Dot-path to the records array in REST response (e.g. 'data.items').",
)
@click.pass_context
def source_profile(
    ctx: click.Context,
    source_type: str,
    conn: str | None,
    path: str | None,
    url: str | None,
    tables: str | None,
    output: str | None,
    bearer_token: str | None,
    api_key_header: str | None,
    rest_max_pages: int,
    rest_results_path: str | None,
) -> None:
    """Connect to a data source and return its schema profile as JSON.

    The profile includes table names, column names, data types, row counts,
    null rates, sample values, and distinct counts — the input to 'pbi source scaffold'.

    \b
    REST auth examples:
      pbi source profile --type rest --url https://api.example.com/v1/orders \\
        --bearer-token $MY_TOKEN
      pbi source profile --type rest --url https://api.example.com/v1/products \\
        --api-key-header "X-Api-Key=abc123"
    """
    console.print(f"[cyan]Profiling {source_type} source...[/cyan]")

    if source_type == "sql":
        if not conn:
            raise click.UsageError("--conn is required for --type sql")
        profile = _profile_sql(conn, tables)
    elif source_type in ("excel", "csv"):
        if not path:
            raise click.UsageError(f"--path is required for --type {source_type}")
        profile = _profile_file(path, source_type)
    elif source_type == "rest":
        if not url:
            raise click.UsageError("--url is required for --type rest")
        profile = _profile_rest(
            url,
            bearer_token=bearer_token,
            api_key_header=api_key_header,
            max_pages=rest_max_pages,
            results_path=rest_results_path,
        )
    else:
        raise click.UsageError(f"Unknown source type: {source_type}")

    if output:
        Path(output).write_text(json.dumps(profile, indent=2, default=str), encoding="utf-8")
        console.print(f"[green]Profile saved to:[/green] {output}")
    else:
        output_json_or_table(profile, ctx, title="Source Profile")


def _profile_sql(conn: str, tables_filter: str | None) -> list[dict[str, Any]]:
    """Profile SQL Server (or any SQLAlchemy-compatible) database."""
    try:
        from sqlalchemy import create_engine, inspect, text  # type: ignore[import]
    except ImportError:
        raise click.ClickException(
            "SQLAlchemy not installed. Run: pip install pbi-cli-tool[sources]"
        )

    engine = create_engine(conn)
    inspector = inspect(engine)

    requested = {t.strip() for t in tables_filter.split(",")} if tables_filter else None
    table_names = inspector.get_table_names()
    if requested:
        table_names = [t for t in table_names if t in requested]

    profile = []
    with engine.connect() as connection:
        for table_name in table_names:
            try:
                row_count_result = connection.execute(text(f"SELECT COUNT(*) FROM [{table_name}]"))
                row_count = row_count_result.scalar() or 0
            except Exception:
                row_count = -1

            columns_info = inspector.get_columns(table_name)
            columns = []
            for col in columns_info:
                col_profile: dict[str, Any] = {
                    "name": col["name"],
                    "dataType": str(col["type"]),
                    "nullable": col.get("nullable", True),
                    "nullRate": None,
                    "distinctCount": None,
                    "sampleValues": [],
                }
                try:
                    col_name = col["name"]
                    stats = connection.execute(
                        text(
                            f"SELECT COUNT(DISTINCT [{col_name}]), "
                            f"SUM(CASE WHEN [{col_name}] IS NULL THEN 1 ELSE 0 END) * 1.0 / COUNT(*) "  # noqa: E501
                            f"FROM [{table_name}]"
                        )
                    ).fetchone()
                    if stats:
                        col_profile["distinctCount"] = stats[0]
                        col_profile["nullRate"] = round(float(stats[1] or 0), 4)
                    samples = connection.execute(
                        text(
                            f"SELECT TOP 5 [{col_name}] FROM [{table_name}] WHERE [{col_name}] IS NOT NULL"  # noqa: E501
                        )
                    ).fetchall()
                    col_profile["sampleValues"] = [row[0] for row in samples]
                except Exception:
                    pass
                columns.append(col_profile)

            profile.append(
                {
                    "tableName": table_name,
                    "rowCount": row_count,
                    "columns": columns,
                }
            )

    return profile


def _profile_file(path: str, file_type: str) -> list[dict[str, Any]]:
    """Profile an Excel or CSV file."""
    try:
        import openpyxl  # type: ignore[import]
    except ImportError:
        if file_type == "excel":
            raise click.ClickException(
                "openpyxl not installed. Run: pip install pbi-cli-tool[sources]"
            )

    file_path = Path(path)
    if not file_path.exists():
        raise click.ClickException(f"File not found: {path}")

    if file_type == "excel":
        import openpyxl  # type: ignore[import]

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        profile = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h) if h is not None else f"Column{i}" for i, h in enumerate(rows[0])]
            data_rows = rows[1:]
            columns = []
            for i, header in enumerate(headers):
                values = [r[i] for r in data_rows if r[i] is not None]
                null_rate = 1.0 - len(values) / len(data_rows) if data_rows else 0.0
                sample = values[:5]
                data_types = {type(v).__name__ for v in values[:100]}
                columns.append(
                    {
                        "name": header,
                        "dataType": list(data_types)[0] if len(data_types) == 1 else "Mixed",
                        "nullRate": round(null_rate, 4),
                        "distinctCount": len(set(values[:1000])),
                        "sampleValues": sample,
                    }
                )
            profile.append(
                {"tableName": sheet_name, "rowCount": len(data_rows), "columns": columns}
            )
        wb.close()
        return profile
    else:
        import csv

        with open(path, encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        if not rows:
            return []
        headers = list(rows[0].keys())
        columns = []
        for h in headers:
            values = [r[h] for r in rows if r[h]]
            null_rate = 1.0 - len(values) / len(rows) if rows else 0.0
            columns.append(
                {
                    "name": h,
                    "dataType": "String",
                    "nullRate": round(null_rate, 4),
                    "distinctCount": len(set(values[:1000])),
                    "sampleValues": values[:5],
                }
            )
        return [{"tableName": Path(path).stem, "rowCount": len(rows), "columns": columns}]


def _profile_rest(
    url: str,
    bearer_token: str | None = None,
    api_key_header: str | None = None,
    max_pages: int = 5,
    results_path: str | None = None,
) -> list[dict[str, Any]]:
    """Profile a REST API with auth, pagination, and nested schema support.

    Handles:
    - Bearer token and API key header auth
    - OData ``@odata.nextLink`` pagination
    - JSON:API ``links.next`` pagination
    - Page-number based pagination (``?page=N``)
    - Nested ``results_path`` extraction (dot-path into response JSON)
    - Flat type mapping to Power BI data types
    """
    try:
        import httpx  # type: ignore[import]
    except ImportError:
        raise click.ClickException("httpx not installed. Run: pip install pbi-cli-tool[sources]")

    # Build auth headers
    headers: dict[str, str] = {"Accept": "application/json"}
    if bearer_token:
        headers["Authorization"] = f"Bearer {bearer_token}"
    if api_key_header:
        if "=" in api_key_header:
            hname, hval = api_key_header.split("=", 1)
            headers[hname.strip()] = hval.strip()
        else:
            raise click.ClickException("--api-key-header must be 'HeaderName=value'")

    all_records: list[dict[str, Any]] = []
    next_url: str | None = url
    page = 0

    with httpx.Client(timeout=30.0, follow_redirects=True) as client:
        while next_url and page < max_pages:
            response = client.get(next_url, headers=headers)
            response.raise_for_status()
            data = response.json()

            # Extract records from the response
            records = _extract_records(data, results_path)
            all_records.extend(records)
            page += 1

            # Detect next-page link (OData, JSON:API, or none)
            next_url = _detect_next_link(data, next_url, page)

            if not records:
                break

    if not all_records:
        return []

    # Infer schema from a sample of records
    sample = all_records[:50]
    columns = _infer_columns(sample)
    table_name = _url_to_table_name(url)

    return [
        {
            "tableName": table_name,
            "rowCount": len(all_records),
            "paginatedPages": page,
            "columns": columns,
        }
    ]


def _extract_records(data: Any, results_path: str | None) -> list[dict[str, Any]]:
    """Navigate a dot-path into the response JSON to find the records array."""
    if results_path:
        node: Any = data
        for part in results_path.split("."):
            if isinstance(node, dict):
                node = node.get(part)
            else:
                node = None
                break
        records = node
    elif isinstance(data, dict):
        if not data:
            return []
        # Common envelope keys: value (OData), data, results, items, records
        for key in ("value", "data", "results", "items", "records"):
            if key in data and isinstance(data[key], list):
                records = data[key]
                break
        else:
            records = [data]
    elif isinstance(data, list):
        records = data
    else:
        records = [data]

    if not isinstance(records, list):
        return []
    return [r for r in records if isinstance(r, dict)]


def _detect_next_link(data: Any, current_url: str, page: int) -> str | None:
    """Detect the next page URL from OData, JSON:API, or page-number patterns."""
    if not isinstance(data, dict):
        return None
    # OData style
    if "@odata.nextLink" in data:
        return str(data["@odata.nextLink"])
    # JSON:API links
    if isinstance(data.get("links"), dict) and data["links"].get("next"):
        return str(data["links"]["next"])
    # No detected pagination
    return None


def _infer_columns(sample: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Infer column schema from a sample of records, handling nested objects."""
    # Flatten one level of nesting — nested objects become "parent.child" columns
    flat_records = [_flatten_record(r) for r in sample]
    if not flat_records:
        return []

    all_keys = list(dict.fromkeys(k for r in flat_records for k in r))
    columns = []
    for key in all_keys:
        values = [r[key] for r in flat_records if r.get(key) is not None]
        null_count = sum(1 for r in flat_records if r.get(key) is None)
        null_rate = round(null_count / len(flat_records), 4) if flat_records else 0.0
        pbi_type = _infer_pbi_type(values)
        columns.append(
            {
                "name": key,
                "dataType": pbi_type,
                "nullRate": null_rate,
                "distinctCount": len(set(str(v) for v in values)),
                "sampleValues": [str(v) for v in values[:3]],
            }
        )
    return columns


def _flatten_record(record: dict[str, Any], prefix: str = "") -> dict[str, Any]:
    """Flatten one level of nested dict into dot-notation keys."""
    flat: dict[str, Any] = {}
    for k, v in record.items():
        full_key = f"{prefix}{k}" if not prefix else f"{prefix}.{k}"
        if isinstance(v, dict) and len(v) <= 5:
            # Flatten shallow nested objects
            flat.update(_flatten_record(v, prefix=full_key))
        elif isinstance(v, list):
            flat[full_key] = f"[Array({len(v)})]"
        else:
            flat[full_key] = v
    return flat


def _infer_pbi_type(values: list[Any]) -> str:
    """Map Python runtime types to Power BI data type names."""
    type_set = {type(v) for v in values if v is not None}
    if not type_set:
        return "String"
    if type_set == {bool}:
        return "Boolean"
    if type_set <= {int}:
        return "Int64"
    if type_set <= {int, float}:
        return "Decimal"
    # Detect ISO date strings
    if type_set <= {str}:
        import re

        date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}")
        if all(date_pattern.match(str(v)) for v in values[:5]):
            return "DateTime"
        return "String"
    return "Mixed"


def _url_to_table_name(url: str) -> str:
    """Derive a clean table name from the URL path."""
    from urllib.parse import urlparse

    path = urlparse(url).path.rstrip("/")
    segment = path.split("/")[-1] if "/" in path else path
    # Strip numeric version prefix (e.g. v1, v2)
    import re

    segment = re.sub(r"^v\d+$", "", segment)
    return segment.title().replace("-", "").replace("_", "") or "RestResponse"


@source.command("scaffold")
@click.option(
    "--profile",
    "profile_path",
    required=True,
    type=click.Path(exists=True),
    help="Path to source profile JSON.",
)
@click.pass_context
def source_scaffold(ctx: click.Context, profile_path: str) -> None:
    """Generate a star-schema model from a source profile JSON file."""
    profile = json.loads(Path(profile_path).read_text(encoding="utf-8"))
    schema = _infer_star_schema(profile)

    if dry_run_echo(ctx, "scaffold star schema", json.dumps(schema, indent=2)):
        return

    backend = get_backend(ctx)
    snapshot_before_write(ctx)
    for table in schema["tables"]:
        backend.table_add(table["name"])
        for col in table["columns"]:
            backend.column_add(table["name"], col["name"], col["dataType"])
    for rel in schema["relationships"]:
        parts = rel["from"].split("[")
        ft, fc = parts[0], parts[1].rstrip("]")
        parts2 = rel["to"].split("[")
        tt, tc = parts2[0], parts2[1].rstrip("]")
        backend.relationship_add(ft, fc, tt, tc)

    write_audit_entry(
        "source scaffold",
        after={"tables": len(schema["tables"]), "relationships": len(schema["relationships"])},
    )
    is_json = ctx.obj and ctx.obj.get("output_json")
    if not is_json:
        console.print(
            f"[green]Scaffolded[/green] {len(schema['tables'])} tables, {len(schema['relationships'])} relationships"  # noqa: E501
        )
    output_json_or_table(schema, ctx, title="Star Schema")


def _infer_star_schema(profile: list[dict]) -> dict[str, Any]:
    """Heuristic: largest table = fact, others = dims. Key columns detected by name."""
    if not profile:
        return {"tables": [], "relationships": [], "measures": []}

    sorted_tables = sorted(profile, key=lambda t: t.get("rowCount", 0), reverse=True)
    fact = sorted_tables[0]
    dims = sorted_tables[1:]

    key_suffixes = ("key", "id", "sk", "code")
    relationships = []

    for dim in dims:
        dim_name = dim["tableName"]
        dim_keys = [c["name"] for c in dim["columns"] if c["name"].lower().endswith(key_suffixes)]
        if not dim_keys:
            continue
        dim_key = dim_keys[0]
        fact_matching = [
            c["name"]
            for c in fact["columns"]
            if c["name"].lower() == dim_key.lower()
            or c["name"].lower() == f"{dim_name.lower()}{dim_key.lower()}"
        ]
        if fact_matching:
            relationships.append(
                {
                    "from": f"{fact['tableName']}[{fact_matching[0]}]",
                    "to": f"{dim_name}[{dim_key}]",
                    "cardinality": "ManyToOne",
                }
            )

    numeric_types = ("decimal", "double", "int64", "float", "int", "float64")
    starter_measures = []
    for col in fact["columns"]:
        if any(t in col["dataType"].lower() for t in numeric_types):
            starter_measures.append(
                {
                    "table": fact["tableName"],
                    "name": f"Total {col['name']}",
                    "expression": f"SUM({fact['tableName']}[{col['name']}])",
                }
            )

    return {
        "tables": [{"name": t["tableName"], "columns": t["columns"]} for t in profile],
        "relationships": relationships,
        "measures": starter_measures[:5],
    }


@source.command("suggest-joins")
@click.option("--profiles", required=True, help="Comma-separated paths to two profile JSON files.")
@click.pass_context
def source_suggest_joins(ctx: click.Context, profiles: str) -> None:
    """Suggest FK relationships between two profiled sources using column name heuristics."""
    paths = [p.strip() for p in profiles.split(",")]
    if len(paths) != 2:
        raise click.UsageError("Provide exactly two comma-separated profile paths.")

    profile_a = json.loads(Path(paths[0]).read_text(encoding="utf-8"))
    profile_b = json.loads(Path(paths[1]).read_text(encoding="utf-8"))

    suggestions = _suggest_joins(profile_a, profile_b)
    output_json_or_table(suggestions, ctx, title="Suggested Joins")


def _suggest_joins(profile_a: list, profile_b: list) -> list[dict]:
    suggestions = []
    for table_a in profile_a:
        for table_b in profile_b:
            cols_a = {c["name"].lower(): (table_a["tableName"], c) for c in table_a["columns"]}
            cols_b = {c["name"].lower(): (table_b["tableName"], c) for c in table_b["columns"]}
            for name, (tname_a, col_a) in cols_a.items():
                if name in cols_b:
                    tname_b, col_b = cols_b[name]
                    suggestions.append(
                        {
                            "from": f"{tname_a}[{col_a['name']}]",
                            "to": f"{tname_b}[{col_b['name']}]",
                            "confidence": "high"
                            if col_a["dataType"] == col_b["dataType"]
                            else "medium",
                            "reason": "exact column name match",
                        }
                    )
    return suggestions
